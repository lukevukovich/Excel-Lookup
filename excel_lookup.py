from openpyxl import load_workbook


# Get file name from user
def get_workbook():
    found = False
    wb = None
    # Loop until file is found
    while not found:
        file_name = input("Excel File: ")

        # If file is not found/invalid
        try:
            wb = load_workbook(file_name)
            found = True
        except FileNotFoundError:
            print("File Not Found\n")
        except (Exception,):
            print("Invalid File\n")

    print("Excel File Found")
    return wb


# Print list of all data columns
def print_search_cols(ws):
    print("\nSearch Fields:")
    end_col = 65 + len(ws[1])

    # Loop for as many columns of data there are
    col_counter = 65
    counter = 1
    while col_counter < end_col:
        title = str(chr(col_counter)) + str(1)
        # List all columns
        print(str(counter) + ": " + str(ws[title].value))
        col_counter += 1
        counter += 1


# Get search column input
def get_search_col(ws):
    end_col = 65 + len(ws[1])

    # Get input for search column
    good_val = False
    col = None
    while not good_val:
        try:
            col = input("\nSelect Search Field: ")
            # Enter x to quit
            if col.lower() == 'x':
                # Not a search field, used as sentinel value
                col = 0
            else:
                col = int(col)

            if col >= 0 and col <= end_col - 65:
                good_val = True
                col -= 1
            else:
                print("Invalid Field")
        except ValueError:
            print("Invalid Field")
        except TypeError:
            print("Invalid Field")

    return col


def get_data_arrays(ws, search_col):
    num_data = len(ws["A"]) - 1

    # Create arrays of data
    # Parallel arrays to keep track of index in sheet
    data = [None] * num_data
    index = [None] * num_data

    # Set data of arrays to search column data
    for (i) in range(num_data):
        data[i] = str(ws[search_col + str(i + 1)].value)
        index[i] = i + 1

    return data, index


# Selection sort based on data_array data
def sort(data_array, index_array):
    for step in range(1, len(data_array)):
        key1 = data_array[step]
        key2 = index_array[step]
        j = step - 1

        while j >= 0 and key1 < data_array[j]:
            data_array[j + 1] = data_array[j]
            index_array[j + 1] = index_array[j]
            j = j - 1

        data_array[j + 1] = key1
        index_array[j + 1] = key2

    return data_array, index_array


# Find data using binary search
def search(data_search, data_array, index_array):
    # Binary search
    low = 0
    high = len(data_array) - 1

    while low <= high:
        mid = (high + low) // 2

        if str(data_array[mid]).lower() < data_search.lower():
            low = mid + 1

        elif str(data_array[mid]).lower() > data_search.lower():
            high = mid - 1

        else:
            return index_array[mid]

    return -1


# Print employee data
def print_data(ws, row):
    end_col = 65 + len(ws[1])

    # Loop for as many columns of data there are
    counter = 65
    while counter < end_col:
        title = str(chr(counter)) + str(1)
        index = str(chr(counter)) + str(row)
        # List data
        print(str(ws[title].value) + ": " + str(ws[index].value))
        counter += 1


def main():
    print("Excel Data Lookup\n")

    # Connect to excel file
    wb = get_workbook()
    print("\nConnected to File")

    ws = wb.active

    # Find number of data entries
    num_data = len(ws['A']) - 1
    print("Number of Data Entries: " + str(num_data))

    # Column select loop
    run_select = True
    while run_select:
        # Get search column input
        print_search_cols(ws)
        search_col = get_search_col(ws)

        if search_col < 0:
            run_select = False
        else:
            search_col = str(chr(65 + search_col))

            # Get title of search column
            a_title = ws[search_col + "1"].value

            # Get array of data & index and sort array
            # Used for searching
            d, i = get_data_arrays(ws, search_col)
            data, index = sort(d, i)

            # Search loop
            run_search = True
            while run_search:
                data_search = input("\nEnter " + str(a_title) + ": ")

                if data_search.lower() == 'x':
                    run_search = False
                else:
                    result = search(data_search, data, index)
                    if result >= 0:
                        print_data(ws, result)
                    else:
                        print("Data Not Found")


main()
